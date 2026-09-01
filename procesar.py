# -*- coding: utf-8 -*-
"""
TRACTOCAR · Control de Ventas COMEX
-----------------------------------
Se conecta a la carpeta (y subcarpetas) con los archivos IMPO/EXPO/NACIONAL/CEDIS,
hace toda la limpieza y la unión, y genera:
    - TRACTOCAR_UNIFICADO.xlsx   (la unión limpia, lista para Power BI/Excel)
    - dashboard.html             (tablero interactivo; se abre solo en el navegador)

USO (en la terminal de VS Code):
    pip install pandas openpyxl
    python procesar.py
"""

import os, sys, json, glob, webbrowser, unicodedata, warnings, datetime as dt
import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")  # oculta avisos técnicos (no son errores) para una salida limpia

# ====================== CONFIGURA AQUÍ TU CARPETA ======================
# Pega la ruta de la carpeta ARCHIVOS (la que contiene CEDIS, EXPO, IMPU, NACIONAL).
CARPETA = r"C:\Users\jarias\OneDrive - TRACTOCAR LOGISTICS SAS\POWER BI JEFFER\ARCHIVOS"
# Carpeta del mes actual (cedis2, expo2, impu2) — misma estructura, se lee igual
CARPETA2 = r"C:\Users\jarias\OneDrive - TRACTOCAR LOGISTICS SAS\POWER BI JEFFER\ACHIVOS 2"
# =======================================================================

# ---------- Cruce Samaritima (opcional) ----------
# Ruta del archivo 'exportación Ajover.xlsx' (hoja 'VACIOS AJOVER'). Si no existe, se omite el cruce.
ARCHIVO_AJOVER = r"C:\Users\jarias\OneDrive - TRACTOCAR LOGISTICS SAS\Archivos de Data Quality Analyst Tractocar - Analisis Operacion y Venta\14.Comex\EXPO\exportación Ajover.xlsx"
HOJA_AJOVER = "VACIOS AJOVER"
ARCHIVO_AJCOMEX = r"C:\Users\jarias\OneDrive - TRACTOCAR LOGISTICS SAS\Archivos de Data Quality Analyst Tractocar - Analisis Operacion y Venta\14.comex\IMPO\IMPORTACIÓN COMEX.xlsx"
PATIOS_SAMARITIMA = ["SIMARITIMA_NUEVO"]
AJUSTE_SAMARITIMA = 340126   # COP a sumar al 'a facturar' por cada manifiesto que tocó Samaritima
# ---------------------------------------------------

# ---------- Enriquecimiento DITAR (desde despachoscomex.tractocar.com) ----------
# Exporta el archivo desde la web app y colócalo junto a procesar.py
ARCHIVO_DITAR = "ditar_despachos.json"
DITAR_NOMBRE_PATRON = "DITAR"       # substring en ClienteNombre para identificar filas DITAR
DITAR_ALERTA_MONTO = 2_500_000      # COP: alerta si AFacturar o APagar supera este valor por fila
# ---------------------------------------------------------------------------------

NIT_PROPIO = "9005033252"                       # TRACTOCAR LOGISTICS SAS
BASE = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------- utilidades
def carpeta_datos():
    if len(sys.argv) > 1 and sys.argv[1].strip():
        return sys.argv[1].strip()
    if os.environ.get("TRACTOCAR_DIR"):
        return os.environ["TRACTOCAR_DIR"]
    if CARPETA and os.path.isdir(CARPETA):
        return CARPETA
    return BASE

def carpetas_datos():
    """Retorna lista de carpetas a escanear (CARPETA + CARPETA2 si existe)."""
    carps = []
    c1 = carpeta_datos()
    if os.path.isdir(c1):
        carps.append(c1)
    if CARPETA2 and os.path.isdir(CARPETA2):
        carps.append(CARPETA2)
    return carps


def _norm(s):
    """Normaliza un nombre de columna: sin tildes, minúsculas, sin espacios de más."""
    s = str(s).replace("\xa0", " ").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return " ".join(s.split())


def norm_nit(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace("-", "").replace(" ", "").replace(".", "")


def split_cuenta(cc):
    """De 'BTA-TN-0022' saca ciudad='BTA', token='TN'. De 'CEDI-BMED-0043' -> token='CED'."""
    if pd.isna(cc) or cc == "":
        return ("", "")
    p = str(cc).strip().split("-")
    if len(p) >= 2:
        if p[0].upper().startswith("CED"):
            return (p[0], "CED")
        return (p[0], p[1].upper())
    return (str(cc), "")


def seg_from(fuente, token):
    if fuente == "IMPO":
        return "Comex Impo"
    if fuente == "EXPO":
        return "Comex Expo"
    if token == "AC":
        return "Alto Cubicaje"
    if token == "CED":
        return "Cedis"
    if token == "TN":
        return "Nacional"
    return f"Nacional ({token})" if token else "Nacional"


def env_num(x):
    """Extrae el número de envío: 'TCL.2250822' -> 2250822 ; 2226138.0 -> 2226138."""
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        try:
            return int(round(float(x)))
        except Exception:
            return None
    s = "".join(ch for ch in str(x) if ch.isdigit())
    return int(s) if s else None


def archivo_ajover():
    if os.environ.get("TRACTOCAR_AJOVER"):
        return os.environ["TRACTOCAR_AJOVER"]
    return ARCHIVO_AJOVER


def _copiar_ajover(ruta):
    """Copia el archivo a una ruta temporal para evitar bloqueo de OneDrive."""
    import shutil, tempfile
    tmp = os.path.join(tempfile.gettempdir(), "ajover_tmp.xlsx")
    shutil.copy2(ruta, tmp)
    return tmp


def leer_ajover_comex(stats):
    """Lee IMPO AJOVER e IMPO CV de IMPORTACIÓN COMEX.xlsx y calcula KPIs de retiros y devoluciones."""
    import shutil, tempfile
    HOY = dt.date.today()
    OBS_EXENTO = {
        "El pedido fue enviado con Bodegajes, TRACTOCAR NO ASUME EXTRACOSTOS",
        "El pedido fue enviado con demoras y bodegajes, TRACTOCAR NO ASUME EXTRACOSTOS",
        "El pedido fue enviado con demoras, TRACTOCAR NO ASUME EXTRACOSTOS",
        "El pedido fue enviado el mismo día que tenia Bodegajes, TRACTOCAR NO ASUME EXTRACOSTOS",
        "Las planillas se reflearon el mismo día que vencian los bodegajes, TRACTOCAR NO ASUME EXTRACOSTOS",
    }
    ruta = os.environ.get("TRACTOCAR_AJCOMEX", ARCHIVO_AJCOMEX)
    if not os.path.isfile(ruta):
        stats["avisos"].append(f"Ajover COMEX omitido: no se encontró '{ruta}'")
        return {}
    try:
        tmp = os.path.join(tempfile.gettempdir(), "ajcomex_tmp.xlsx")
        shutil.copy2(ruta, tmp)
    except Exception as e:
        stats["avisos"].append(f"Ajover COMEX: no pude copiar archivo ({e})")
        return {}

    def _ts(df, col):
        return pd.to_datetime(df[col], errors="coerce", dayfirst=True) if col and col in df.columns else pd.Series([pd.NaT]*len(df))

    def _str(df, col):
        if col not in df.columns: return pd.Series([""]*len(df))
        return df[col].astype(str).str.strip().replace({"nan":"","None":"","NAN":"","NONE":""})

    result = {}

    for hoja, fuente in [("IMPO AJOVER","AJOVER"), ("IMPO CV","CV")]:
        try:
            df = pd.read_excel(tmp, sheet_name=hoja)
            df.columns = [str(c).strip() for c in df.columns]
            nm = {_norm(c): c for c in df.columns}

            def _col(*keys):
                for k in keys:
                    if k in nm: return nm[k]
                return None

            c_cliente   = _col("cliente")
            c_cont      = _col("contenedor")
            c_tamano    = _col("tamano", "tamaño")
            c_eta       = _col("eta")
            c_pedido    = _col("pedido")
            c_do        = _col("do")
            c_linea     = _col("linea")
            c_terminal  = _col("terminal portuaria (retiro)", "lugar de retiro")
            c_destino   = _col("destino descargue")
            c_tipo_dev  = _col("tipo de devolucion")
            c_bodegaje  = _col("fecha de bodegaje")
            c_max_dev   = _col("fecha maxima de devolucion de unidad vacia (demoras)")
            c_cita_ret  = _col("fecha y hora de cita de retiro del contenedor", "cita")
            c_cita_ret_repr = _col("fecha y hora de cita reprogramada de retiro del contenedor")
            c_llegada_ret   = _col("fecha y hora de llegada a retiro del contenedor")
            c_salida_puer   = _col("fecha y hora de salida de puerto")
            c_llegada_desc  = _col("fecha y hora de llegada a descargue")
            c_cita_dev      = _col("fecha y hora de cita de devolucion unidad vacia")
            c_llegada_dev   = _col("fecha y hora de llegada a devolucion unidad vacia")
            c_lugar_dev     = _col("lugar devolucion unidad vacia", "sitio devolucion")
            c_cierre        = _col("cierre de pedido")
            c_fecrecib      = _col("fecha de recibido del pedido", "fecha recibido del pedido", "fecha recibido", "recibido pedido")
            c_salida_desc   = _col("fecha y hora de salida a descargue")
            c_llegada_patio = _col("fecha y hora de llegada a patio temporal")
            c_salida_patio  = _col("fecha y hora de salida patio temporal")

            f_eta         = _ts(df, c_eta)
            f_bodegaje    = _ts(df, c_bodegaje)
            f_max_dev     = _ts(df, c_max_dev)
            f_cita_ret    = _ts(df, c_cita_ret)
            f_cita_ret_repr = _ts(df, c_cita_ret_repr)
            f_llegada_ret = _ts(df, c_llegada_ret)
            f_salida_puer = _ts(df, c_salida_puer)
            f_llegada_desc= _ts(df, c_llegada_desc)
            f_fecrecib    = _ts(df, c_fecrecib)
            f_salida_desc = _ts(df, c_salida_desc)
            f_llegada_patio = _ts(df, c_llegada_patio)
            f_salida_patio  = _ts(df, c_salida_patio)
            f_cita_dev    = _ts(df, c_cita_dev)
            f_llegada_dev = _ts(df, c_llegada_dev)
            obs_series    = _str(df, _col("observacion", "observaciones"))

            rows = []
            tend_raw = {}  # {YYYY-MM: {...}}

            for i in range(len(df)):
                fecrecib= f_fecrecib.iloc[i]
                eta     = f_eta.iloc[i]
                bod     = f_bodegaje.iloc[i]
                max_dev = f_max_dev.iloc[i]
                cita_r  = f_cita_ret.iloc[i]
                cita_rr = f_cita_ret_repr.iloc[i]
                llegr   = f_llegada_ret.iloc[i]
                salp    = f_salida_puer.iloc[i]
                llegdc  = f_llegada_desc.iloc[i]
                sald    = f_salida_desc.iloc[i]
                llegp   = f_llegada_patio.iloc[i]
                salp2   = f_salida_patio.iloc[i]
                cita_d  = f_cita_dev.iloc[i]
                llegdv  = f_llegada_dev.iloc[i]
                obs_val = obs_series.iloc[i]

                mes_iso  = eta.strftime("%Y-%m") if not pd.isna(eta) else ""
                mes_rec  = fecrecib.strftime("%Y-%m") if not pd.isna(fecrecib) else ""

                # ── cumplimiento retiro ──────────────────────────────────────
                if fuente == "CV":
                    # CV: compara llegada real (llegr) vs cita (cita_r)
                    if obs_val in OBS_EXENTO:
                        cumpl_ret = "Exento (no asume)"
                    elif pd.isna(cita_r):
                        cumpl_ret = "Sin cita"
                    elif pd.isna(llegr):
                        if cita_r >= pd.Timestamp(HOY):
                            cumpl_ret = "Pendiente"
                        else:
                            cumpl_ret = "Sin retiro (vencido)"
                    elif llegr <= cita_r:
                        cumpl_ret = "A tiempo"
                    else:
                        horas = round((llegr - cita_r).total_seconds() / 3600, 1)
                        cumpl_ret = f"Tarde +{horas}h"
                else:
                    # AJOVER: compara cita (cita_r) vs fecha bodegaje (bod)
                    if obs_val in OBS_EXENTO:
                        cumpl_ret = "Exento (no asume)"
                    elif pd.isna(cita_r):
                        if pd.isna(bod):
                            cumpl_ret = "Sin fecha"
                        elif bod.date() >= HOY:
                            cumpl_ret = "Pendiente"
                        else:
                            cumpl_ret = "Sin cita (vencido)"
                    elif not pd.isna(bod) and cita_r.date() > bod.date():
                        dias = (cita_r.date() - bod.date()).days
                        cumpl_ret = f"Tarde +{dias}d"
                    else:
                        cumpl_ret = "A tiempo"

                # ── estado retiro ────────────────────────────────────────────
                if fuente == "CV":
                    # CV: basado en llegr vs cita_r
                    if pd.isna(cita_r):
                        estado_ret = "SIN RETIRO - A TIEMPO"
                    elif pd.isna(llegr):
                        if cita_r >= pd.Timestamp(HOY):
                            estado_ret = "SIN RETIRO - A TIEMPO"
                        else:
                            estado_ret = "SIN RETIRO - ATRASADO"
                    elif llegr <= cita_r:
                        estado_ret = "RETIRADO - A TIEMPO"
                    else:
                        estado_ret = "RETIRADO - ATRASADO"
                else:
                    # AJOVER: basado en cita_r vs bod
                    if pd.isna(cita_r):
                        if pd.isna(bod) or bod.date() >= HOY:
                            estado_ret = "SIN RETIRO - A TIEMPO"
                        else:
                            estado_ret = "SIN RETIRO - ATRASADO"
                    elif not pd.isna(llegr) and not pd.isna(bod) and cita_r.date() > bod.date():
                        estado_ret = "RETIRADO - ATRASADO"
                    else:
                        estado_ret = "RETIRADO - A TIEMPO"

                # ── cumplimiento devolución (lógica Power Query) ─────────────
                OBS_EXENTO_DEV = "El pedido fue enviado con demoras, TRACTOCAR NO ASUME EXTRACOSTOS"
                if pd.isna(max_dev):
                    cumpl_dev = "A tiempo (sin límite)"
                elif obs_val == OBS_EXENTO_DEV:
                    cumpl_dev = "Exento (no asume)"
                elif pd.isna(llegdv):
                    if max_dev.date() >= HOY:
                        cumpl_dev = "Pendiente"
                    else:
                        cumpl_dev = "Sin devolución (vencido)"
                elif llegdv.date() > max_dev.date():
                    dias = (llegdv.date() - max_dev.date()).days
                    cumpl_dev = f"Tarde +{dias}d"
                else:
                    cumpl_dev = "A tiempo"

                # ── estado devolución ────────────────────────────────────────
                if pd.isna(llegdv) and not pd.isna(max_dev) and max_dev.date() >= HOY:
                    estado_dev = "SIN DEVOLUCION - A TIEMPO"
                elif pd.isna(max_dev) and not pd.isna(llegdv):
                    estado_dev = "CON DEVOLUCION - A TIEMPO"
                elif pd.isna(llegdv) and not pd.isna(max_dev) and max_dev.date() < HOY:
                    estado_dev = "SIN DEVOLUCION - ATRASADO"
                elif not pd.isna(llegdv) and not pd.isna(max_dev) and llegdv.date() > max_dev.date():
                    estado_dev = "CON DEVOLUCION - ATRASADO"
                else:
                    estado_dev = "CON DEVOLUCION - A TIEMPO"

                # ── tiempos (horas) ──────────────────────────────────────────
                def _dh(a, b):
                    if pd.isna(a) or pd.isna(b): return None
                    h = (b - a).total_seconds() / 3600
                    return round(h, 2)

                # En ruta: solo si llegada y salida son el mismo día
                if not pd.isna(salp) and not pd.isna(llegdc) and salp.date() == llegdc.date():
                    t_en_ruta = _dh(salp, llegdc)
                else:
                    t_en_ruta = None
                t_descargando   = _dh(llegdc, sald)
                t_prom_dev      = _dh(sald, llegdv)
                t_salida_vs_cita= _dh(cita_r, salp)
                t_en_ruta2      = _dh(salp2 if not pd.isna(salp2) else sald, llegdv)
                t_patio_temp    = _dh(llegp, salp2) if not pd.isna(salp2) else None
                t_espera_ret    = _dh(cita_r, llegr)
                t_en_puerto     = _dh(llegr, salp)

                # ── estado contenedor ────────────────────────────────────────
                if pd.isna(cita_r):
                    estado_cont = "SIN LLEGADA DE RETIRO DEL CONTENEDOR"
                elif pd.isna(salp):
                    estado_cont = "SIN SALIDA DE PUERTO"
                elif pd.isna(llegdc):
                    estado_cont = "EN RUTA"
                elif pd.isna(sald):
                    estado_cont = "DESCARGANDO"
                elif pd.isna(llegp) and not pd.isna(llegdv):
                    estado_cont = "FINALIZADO - DEVOLUCION"
                elif pd.isna(llegp) and pd.isna(llegdv):
                    estado_cont = "DESCARGADO"
                elif not pd.isna(llegp) and pd.isna(llegdv):
                    estado_cont = "BAJADO EN PATIO TEMPORAL"
                elif not pd.isna(llegp) and not pd.isna(llegdv):
                    estado_cont = "FINALIZADO - DEVOLUCION"
                elif pd.isna(salp2):
                    estado_cont = "DESCARGADO EN PATIO TEMPORAL"
                elif not pd.isna(llegdv):
                    estado_cont = "FINALIZADO - DEVOLUCION"
                else:
                    estado_cont = "revisar fechas"

                # ── tendencia mes ────────────────────────────────────────────
                if mes_iso:
                    t = tend_raw.setdefault(mes_iso, {
                        "total":0,"ret_aT":0,"ret_tard":0,"ret_pend":0,
                        "dev_aT":0,"dev_tard":0,"dev_pend":0,
                        "t_espera_ret":[],"t_en_puerto":[],"t_en_ruta":[],
                        "t_descargando":[],"t_prom_dev":[]
                    })
                    t["total"] += 1
                    if cumpl_ret in ("A tiempo","Exento (no asume)"): t["ret_aT"] += 1
                    elif "Tarde" in cumpl_ret or "vencido" in cumpl_ret: t["ret_tard"] += 1
                    else:                          t["ret_pend"] += 1
                    if cumpl_dev in ("A tiempo","Exento (no asume)","A tiempo (sin límite)"): t["dev_aT"] += 1
                    elif "Tarde" in cumpl_dev or "vencido" in cumpl_dev: t["dev_tard"] += 1
                    else:                          t["dev_pend"] += 1
                    if t_espera_ret  is not None: t["t_espera_ret"].append(t_espera_ret)
                    if t_en_puerto   is not None: t["t_en_puerto"].append(t_en_puerto)
                    if t_en_ruta     is not None: t["t_en_ruta"].append(t_en_ruta)
                    if t_descargando is not None: t["t_descargando"].append(t_descargando)
                    if t_prom_dev    is not None: t["t_prom_dev"].append(t_prom_dev)

                tam_val  = _str(df, c_tamano).iloc[i]
                ped_val  = _str(df, c_pedido).iloc[i]
                _cs_kw   = ("pallet","plt","carga suelta","suelto","loose","carga suel")
                _is_cs   = any(k in str(tam_val).lower() for k in _cs_kw) or \
                           any(k in str(ped_val).lower()  for k in _cs_kw)
                tipo_carga = "CARGA SUELTA" if _is_cs else "CONTENEDOR"

                rows.append({
                    "fuente":    fuente,
                    "cliente":   _str(df, c_cliente).iloc[i],
                    "cont":      _str(df, c_cont).iloc[i],
                    "tamano":    tam_val,
                    "pedido":    ped_val,
                    "tipo_carga": tipo_carga,
                    "do":        _str(df, c_do).iloc[i],
                    "linea":     _str(df, c_linea).iloc[i],
                    "terminal":  _str(df, c_terminal).iloc[i],
                    "destino":   _str(df, c_destino).iloc[i],
                    "tipo_dev":  _str(df, c_tipo_dev).iloc[i],
                    "observacion": obs_val,
                    "eta":       eta.strftime("%Y-%m-%d") if not pd.isna(eta) else "",
                    "fecrecib":  fecrecib.strftime("%Y-%m-%d") if not pd.isna(fecrecib) else "",
                    "mes_iso":   mes_iso,
                    "mes_rec":   mes_rec,
                    "f_bodegaje":bod.strftime("%Y-%m-%d")         if not pd.isna(bod)    else "",
                    "f_max_dev": max_dev.strftime("%Y-%m-%d")     if not pd.isna(max_dev) else "",
                    "f_cita_ret":cita_r.strftime("%d-%m-%Y %H:%M") if not pd.isna(cita_r) else "",
                    "f_cita_rr": cita_rr.strftime("%d-%m-%Y %H:%M") if not pd.isna(cita_rr) else "",
                    "f_llegr":   llegr.strftime("%d-%m-%Y %H:%M")  if not pd.isna(llegr)  else "",
                    "f_salp":    salp.strftime("%d-%m-%Y %H:%M")   if not pd.isna(salp)   else "",
                    "f_llegdc":  llegdc.strftime("%d-%m-%Y %H:%M") if not pd.isna(llegdc) else "",
                    "f_sald":    sald.strftime("%d-%m-%Y %H:%M")   if not pd.isna(sald)   else "",
                    "f_llegp":   llegp.strftime("%d-%m-%Y %H:%M")  if not pd.isna(llegp)  else "",
                    "f_salp2":   salp2.strftime("%d-%m-%Y %H:%M")  if not pd.isna(salp2)  else "",
                    "f_cita_dev":cita_d.strftime("%d-%m-%Y %H:%M") if not pd.isna(cita_d) else "",
                    "f_llegdv":  llegdv.strftime("%d-%m-%Y %H:%M") if not pd.isna(llegdv) else "",
                    "f_lugar_dev": _str(df, c_lugar_dev).iloc[i],
                    "cierre":    _str(df, c_cierre).iloc[i],
                    "cumpl_ret":  cumpl_ret,
                    "estado_ret": estado_ret,
                    "cumpl_dev":  cumpl_dev,
                    "estado_dev": estado_dev,
                    "estado_cont":estado_cont,
                    "t_espera_ret":     t_espera_ret,
                    "t_en_puerto":      t_en_puerto,
                    "t_en_ruta":        t_en_ruta,
                    "t_descargando":    t_descargando,
                    "t_prom_dev":       t_prom_dev,
                    "t_salida_vs_cita": t_salida_vs_cita,
                    "t_en_ruta2":       t_en_ruta2,
                    "t_patio_temp":     t_patio_temp,
                })

            def _avg(lst): return round(sum(lst)/len(lst),1) if lst else None

            tendencia = []
            for m, v in sorted(tend_raw.items()):
                ret_con = v["ret_aT"] + v["ret_tard"]
                dev_con = v["dev_aT"] + v["dev_tard"]
                tendencia.append({
                    "mes": m,
                    "total": v["total"],
                    "ret_aT": v["ret_aT"], "ret_tard": v["ret_tard"], "ret_pend": v["ret_pend"],
                    "dev_aT": v["dev_aT"], "dev_tard": v["dev_tard"], "dev_pend": v["dev_pend"],
                    "pct_ret": round(v["ret_aT"]/ret_con*100,1) if ret_con else 0,
                    "pct_dev": round(v["dev_aT"]/dev_con*100,1) if dev_con else 0,
                    "avg_espera_ret":  _avg(v["t_espera_ret"]),
                    "avg_en_puerto":   _avg(v["t_en_puerto"]),
                    "avg_en_ruta":     _avg(v["t_en_ruta"]),
                    "avg_descargando": _avg(v["t_descargando"]),
                    "avg_prom_dev":    _avg(v["t_prom_dev"]),
                })

            tot = len(rows)
            ret_aT   = sum(1 for r in rows if r["cumpl_ret"] in ("A tiempo","Exento (no asume)"))
            ret_tard = sum(1 for r in rows if "Tarde" in r["cumpl_ret"] or "vencido" in r["cumpl_ret"])
            ret_pend = sum(1 for r in rows if r["cumpl_ret"] in ("Pendiente","Sin fecha"))
            dev_aT   = sum(1 for r in rows if r["cumpl_dev"] in ("A tiempo","Exento (no asume)","A tiempo (sin límite)"))
            dev_tard = sum(1 for r in rows if "Tarde" in r["cumpl_dev"] or "vencido" in r["cumpl_dev"])
            dev_pend = sum(1 for r in rows if r["cumpl_dev"] == "Pendiente")
            ret_con  = ret_aT + ret_tard
            dev_con  = dev_aT + dev_tard

            result[fuente.lower()] = {
                "total": tot,
                "ret_aT": ret_aT, "ret_tard": ret_tard, "ret_pend": ret_pend,
                "pct_ret": round(ret_aT/ret_con*100,1) if ret_con else 0,
                "dev_aT": dev_aT, "dev_tard": dev_tard, "dev_pend": dev_pend,
                "pct_dev": round(dev_aT/dev_con*100,1) if dev_con else 0,
                "avg_espera_ret":  _avg([r["t_espera_ret"]     for r in rows if r["t_espera_ret"]     is not None]),
                "avg_en_puerto":   _avg([r["t_en_puerto"]       for r in rows if r["t_en_puerto"]       is not None]),
                "avg_en_ruta":     _avg([r["t_en_ruta"]         for r in rows if r["t_en_ruta"]         is not None]),
                "avg_descargando": _avg([r["t_descargando"]     for r in rows if r["t_descargando"]     is not None]),
                "avg_prom_dev":    _avg([r["t_prom_dev"]        for r in rows if r["t_prom_dev"]        is not None]),
                "tendencia": tendencia,
                "rows": rows,
            }
            print(f"  Ajover COMEX {hoja:<20}: {tot} filas | Retiros a tiempo: {ret_aT}/{ret_con} ({result[fuente.lower()]['pct_ret']}%) | Dev a tiempo: {dev_aT}/{dev_con} ({result[fuente.lower()]['pct_dev']}%)")
        except Exception as e:
            import traceback
            stats["avisos"].append(f"Ajover COMEX {hoja} error: {e} | {traceback.format_exc()[:300]}")

    return result


def leer_ajover_completo(stats):
    """Lee VACIOS AJOVER y LLENOS AJOVER y calcula cumplimiento de fechas programadas."""
    ruta = archivo_ajover()
    if not ruta or not os.path.isfile(ruta):
        return None
    try:
        ruta = _copiar_ajover(ruta)
    except Exception as e:
        stats["avisos"].append(f"No se pudo copiar Ajover a temp: {e}")
        return None

    def _col(nmap, *claves):
        for k in claves:
            if k in nmap: return nmap[k]
        return None

    result = {}

    # ── VACIOS AJOVER ──────────────────────────────────────────────
    try:
        va = pd.read_excel(ruta, sheet_name="VACIOS AJOVER")
        va.columns = [str(c).strip() for c in va.columns]
        nm = {_norm(c): c for c in va.columns}

        c_prog     = _col(nm, "fecha y hora recibido de programacion")
        c_aten     = next((v for k, v in nm.items() if "atencion" in k and "retiro" in k), None)
        c_llpatio  = next((v for k, v in nm.items() if "llegada" in k and "vehiculo" in k and "patio" in k), None)
        c_slpatio  = next((v for k, v in nm.items() if "salida" in k and "vehiculo" in k and "patio" in k), None)
        c_llplanta = next((v for k, v in nm.items() if "llegada" in k and "planta" in k and "contenedor" in k), None)
        c_patio    = _col(nm, "patio")
        c_linea    = _col(nm, "linea")
        c_cont     = next((v for k, v in nm.items() if "contenedor" in k and "n" in k[:4]), None)
        c_man      = next((v for k, v in nm.items() if "manifiesto vacio" in k and "radicado" not in k), None)

        def _ts(col):
            return pd.to_datetime(va[col], errors="coerce", dayfirst=True) if col else pd.Series([pd.NaT]*len(va))

        prog_s  = _ts(c_prog)
        aten_s  = _ts(c_aten)
        llpa_s  = _ts(c_llpatio)
        slpa_s  = _ts(c_slpatio)
        llpl_s  = _ts(c_llplanta)

        def _mins(a, b):
            if pd.isna(a) or pd.isna(b) or (b - a).total_seconds() < 0:
                return None
            return round((b - a).total_seconds() / 60, 1)

        a_tiempo = tarde = sin_fecha = 0
        dt1s = []; dt2s = []; dt3s = []; dt4s = []
        va_rows = []
        for i in range(len(va)):
            pg, at = prog_s.iloc[i], aten_s.iloc[i]
            lpa, slpa, llp = llpa_s.iloc[i], slpa_s.iloc[i], llpl_s.iloc[i]
            if pd.isna(pg) or pd.isna(at):
                cumpl = "Sin fecha"; sin_fecha += 1
            elif at.date() <= pg.date():
                cumpl = "A tiempo"; a_tiempo += 1
            else:
                horas = round((at - pg).total_seconds() / 3600, 1)
                cumpl = f"Tarde +{horas}h"; tarde += 1
            d1 = _mins(pg,  at);   d1 and dt1s.append(d1)
            d2 = _mins(at,  lpa);  d2 and dt2s.append(d2)
            d3 = _mins(lpa, slpa); d3 and dt3s.append(d3)
            d4 = _mins(slpa, llp); d4 and dt4s.append(d4)
            va_rows.append({
                "mes":     str(va[_col(nm,"mes")].iloc[i] if _col(nm,"mes") else ""),
                "fprog":   pg.strftime("%d-%m-%Y %H:%M")   if not pd.isna(pg)   else "",
                "faten":   at.strftime("%d-%m-%Y %H:%M")   if not pd.isna(at)   else "",
                "fllpatio":lpa.strftime("%d-%m-%Y %H:%M")  if not pd.isna(lpa)  else "",
                "fslpatio":slpa.strftime("%d-%m-%Y %H:%M") if not pd.isna(slpa) else "",
                "fplanta": llp.strftime("%d-%m-%Y %H:%M")  if not pd.isna(llp)  else "",
                "patio":   str(va[c_patio].iloc[i] if c_patio else ""),
                "linea":   str(va[c_linea].iloc[i] if c_linea else ""),
                "cont":    str(va[c_cont].iloc[i]  if c_cont  else ""),
                "man":     str(va[c_man].iloc[i]   if c_man   else ""),
                "cumpl":   cumpl,
                "dt1": d1, "dt2": d2, "dt3": d3, "dt4": d4,
            })

        def _avg(lst):
            return round(sum(lst)/len(lst), 1) if lst else None

        patio_ct = {}
        if c_patio:
            for v in va[c_patio].dropna():
                k = str(v).strip()
                if k: patio_ct[k] = patio_ct.get(k, 0) + 1
        linea_ct = {}
        if c_linea:
            for v in va[c_linea].dropna():
                k = str(v).strip()
                if k: linea_ct[k] = linea_ct.get(k, 0) + 1

        result["vacios"] = {
            "total": len(va), "a_tiempo": a_tiempo, "tarde": tarde, "sin_fecha": sin_fecha,
            "patios": patio_ct, "lineas": linea_ct, "rows": va_rows,
            "prom": {
                "prog_aten":   _avg(dt1s),
                "aten_llpatio":_avg(dt2s),
                "en_patio":    _avg(dt3s),
                "patio_planta":_avg(dt4s),
            }
        }
    except Exception as e:
        stats["avisos"].append(f"Ajover VACIOS error: {e}"); result["vacios"] = None

    # ── LLENOS AJOVER ──────────────────────────────────────────────
    try:
        ll = pd.read_excel(ruta, sheet_name="LLENOS AJOVER")
        ll.columns = [str(c).strip() for c in ll.columns]
        nm2 = {_norm(c): c for c in ll.columns}

        c_fat     = next((v for k, v in nm2.items() if "fecha" in k and "atencion" in k and "pedido" in k), None) \
                    or next((v for k, v in nm2.items() if "fecha" in k and "atencion" in k), None)
        c_estado  = next((v for k, v in nm2.items() if "estado" in k and "operacion" in k), None)
        c_motivo  = next((v for k, v in nm2.items() if "motivo" in k), None)
        c_planeada= next((v for k, v in nm2.items() if "planeada" in k), None)
        c_llegplant = next((v for k, v in nm2.items() if "llegada" in k and "planta" in k and "planeada" not in k), None)
        c_salida  = next((v for k, v in nm2.items() if "salida" in k and "vehiculo" in k and "cita" in k), None)
        c_cita    = next((v for k, v in nm2.items() if "cita" in k and "puerto" in k and "reprog" not in k), None)
        c_cita_repr = next((v for k, v in nm2.items() if "cita" in k and "puerto" in k and "reprog" in k), None)
        c_motivo_repr = next((v for k, v in nm2.items() if "motivo" in k and "reprog" in k), None)
        c_resp_repr   = next((v for k, v in nm2.items() if "responsable" in k and "reprog" in k), None)
        c_llegpuer= next((v for k, v in nm2.items() if "llegada" in k and "puerto" in k), None)
        c_cont2   = next((v for k, v in nm2.items() if "contenedor" in k), None)
        c_term    = next((v for k, v in nm2.items() if "terminal" in k), None)
        c_placa2  = _col(nm2, "placa")
        c_obs     = next((v for k, v in nm2.items() if "observacion" in k), None)
        c_ob2     = next((v for k, v in nm2.items() if "orden" in k and "base" in k), None) \
                    or next((v for k, v in nm2.items() if k == "ob"), None)
        c_man2    = next((v for k, v in nm2.items() if "manifiesto" in k), None)

        def _ts2(col):
            return pd.to_datetime(ll[col], errors="coerce", dayfirst=True) if col else pd.Series([pd.NaT]*len(ll))

        fat_s   = _ts2(c_fat)
        plan_s  = _ts2(c_planeada)
        llpl_s  = _ts2(c_llegplant)
        sal_s   = _ts2(c_salida)
        cita_s      = _ts2(c_cita)
        cita_repr_s = _ts2(c_cita_repr)
        llpu_s  = _ts2(c_llegpuer)

        # Causas externas: tardanza NO es culpa de Tractocar → cuenta como cumplido
        EXTERNO = ["manifestaci", "accidente de transito", "accidente en la via",
                   "accidente en la vía", "cierre", "paro", "bloqueo", "lluvia",
                   "semaforo", "trafico", "tráfico", "demora en salida de planta",
                   "salida tarde de planta", "no lo dejaron ingresar",
                   "protesta", "orden publica", "orden pública", "huelga",
                   "derrumbe", "via cerrada", "vía cerrada", "represamiento"]
        INTERNO = ["falla mec", "trompo", "mal estado", "averia", "avería",
                   "llanta", "conductor llega tarde", "conductor asignado",
                   "perdida de cita por salida tard", "demora conductor"]

        def _clasif(m):
            ml = _norm(m)
            for k in EXTERNO:
                if k in ml: return "externo"
            for k in INTERNO:
                if k in ml: return "interno"
            return "otro"

        def _dmin(a, b):
            if pd.isna(a) or pd.isna(b): return None
            d = (b - a).total_seconds() / 60
            return None if d < 0 else round(d, 1)

        exitosos = fallidos = 0
        cumpl_cita = no_cumpl = sin_fecha_c = externo_ok = reprog_ajover_ok = 0
        motivos = {}; no_cumpl_motivos = {}; estados = {}; motivos_repr = {}; ll_rows = []
        dt1s = []; dt2s = []; dt3s = []; delta_citas = []
        tendencia_raw = {}  # {YYYY-MM: {total, cumpl, no_cumpl, externo, obs_set, mans}}

        for i in range(len(ll)):
            estado = str(ll[c_estado].iloc[i] if c_estado else "").strip().upper()
            motivo = str(ll[c_motivo].iloc[i] if c_motivo else "").strip()
            obs    = str(ll[c_obs].iloc[i]    if c_obs    else "").strip()
            motivo_repr = str(ll[c_motivo_repr].iloc[i] if c_motivo_repr else "").strip()
            resp_repr   = str(ll[c_resp_repr].iloc[i]   if c_resp_repr   else "").strip()
            if motivo in ("nan","None","NAN","NONE",""): motivo = ""
            if obs    in ("nan","None","NAN","NONE",""): obs = ""
            if motivo_repr in ("nan","None","NAN","NONE",""): motivo_repr = ""
            if resp_repr   in ("nan","None","NAN","NONE",""): resp_repr = ""
            if estado == "EXITOSO": exitosos += 1
            elif estado:
                fallidos += 1
                if motivo: motivos[motivo] = motivos.get(motivo, 0) + 1
            if estado: estados[estado] = estados.get(estado, 0) + 1

            fa  = fat_s.iloc[i]
            pl  = plan_s.iloc[i]
            llp = llpl_s.iloc[i]
            sal = sal_s.iloc[i]
            ct  = cita_s.iloc[i]
            lp  = llpu_s.iloc[i]

            mes_iso = fa.strftime("%Y-%m") if not pd.isna(fa) else ""
            clasif = _clasif(motivo) if motivo else "otro"

            # Cumplimiento cita: ventana de 1 hora
            ct_repr = cita_repr_s.iloc[i]
            resp_es_ajover = "ajover" in resp_repr.lower() if resp_repr else False
            if motivo_repr and resp_repr:
                motivos_repr[motivo_repr] = motivos_repr.get(motivo_repr, 0) + 1
            if pd.isna(ct) or pd.isna(lp):
                cumpl_c = "Sin fecha"; sin_fecha_c += 1
            elif lp <= ct + pd.Timedelta(hours=1):
                cumpl_c = "A tiempo"; cumpl_cita += 1
            elif clasif == "externo":
                mins = round((lp - ct).total_seconds() / 60)
                cumpl_c = f"Tarde +{mins}min (externo)"; externo_ok += 1; cumpl_cita += 1
            elif resp_es_ajover and not pd.isna(ct_repr):
                mins = round((lp - ct).total_seconds() / 60)
                cumpl_c = f"Tarde +{mins}min (reprog Ajover)"; reprog_ajover_ok += 1; cumpl_cita += 1
            else:
                mins = round((lp - ct).total_seconds() / 60)
                cumpl_c = f"Tarde +{mins}min"; no_cumpl += 1
                if motivo: no_cumpl_motivos[motivo] = no_cumpl_motivos.get(motivo, 0) + 1
                else:      no_cumpl_motivos["(sin motivo)"] = no_cumpl_motivos.get("(sin motivo)", 0) + 1

            ob  = str(ll[c_ob2].iloc[i]  if c_ob2  else "").strip()
            man = str(ll[c_man2].iloc[i] if c_man2 else "").strip()
            if ob  in ("nan","None","NAN","NONE",""): ob  = ""
            if man in ("nan","None","NAN","NONE",""): man = ""

            # Tendencia mes a mes (OB y manifiestos siempre; cumplimiento solo con fechas)
            if mes_iso:
                t = tendencia_raw.setdefault(mes_iso, {"total": 0, "cumpl": 0, "no_cumpl": 0, "externo": 0, "obs_set": set(), "mans": 0, "rows_tot": 0})
                t["rows_tot"] += 1
                if ob:  t["obs_set"].add(ob)
                if man: t["mans"] += 1
                if not pd.isna(ct) and not pd.isna(lp):
                    t["total"] += 1
                    if cumpl_c == "A tiempo":
                        t["cumpl"] += 1
                    elif "externo" in cumpl_c:
                        t["cumpl"] += 1; t["externo"] += 1
                    elif "reprog Ajover" in cumpl_c:
                        t["cumpl"] += 1
                    else:
                        t["no_cumpl"] += 1

            # Delta vs cita (en minutos, positivo = tarde)
            if not pd.isna(ct) and not pd.isna(lp):
                delta_citas.append(round((lp - ct).total_seconds() / 60, 1))

            # Tiempos de operación
            d1 = _dmin(pl,  llp)   # planeada → llegada real a planta
            d2 = _dmin(llp, sal)   # llegada planta → salida a cita (cargue)
            d3 = _dmin(sal, lp)    # salida → llegada puerto
            if d1 is not None: dt1s.append(d1)
            if d2 is not None: dt2s.append(d2)
            if d3 is not None: dt3s.append(d3)
            if mes_iso:
                _t = tendencia_raw.get(mes_iso)
                if _t is not None:
                    if d1 is not None: _t.setdefault("dt1s",[]).append(d1)
                    if d2 is not None: _t.setdefault("dt2s",[]).append(d2)
                    if d3 is not None: _t.setdefault("dt3s",[]).append(d3)

            ll_rows.append({
                "fecha":        fa.strftime("%d-%m-%Y")      if not pd.isna(fa)  else "",
                "mes_iso":      mes_iso,
                "ob":           ob,
                "man":          man,
                "cont":         str(ll[c_cont2].iloc[i]  if c_cont2  else ""),
                "terminal":     str(ll[c_term].iloc[i]   if c_term   else ""),
                "placa":        str(ll[c_placa2].iloc[i] if c_placa2 else ""),
                "estado":       estado,
                "motivo":       motivo,
                "clasif":       clasif,
                "obs":          obs,
                "fcita":        ct.strftime("%d-%m-%Y %H:%M")      if not pd.isna(ct)      else "",
                "fcita_repr":   ct_repr.strftime("%d-%m-%Y %H:%M") if not pd.isna(ct_repr) else "",
                "motivo_repr":  motivo_repr,
                "resp_repr":    resp_repr,
                "fllpuerto":    lp.strftime("%d-%m-%Y %H:%M")  if not pd.isna(lp)  else "",
                "fplanta_plan": pl.strftime("%d-%m-%Y %H:%M")  if not pd.isna(pl)  else "",
                "fplanta_real": llp.strftime("%d-%m-%Y %H:%M") if not pd.isna(llp) else "",
                "fsalida":      sal.strftime("%d-%m-%Y %H:%M") if not pd.isna(sal) else "",
                "cumpl_cita":   cumpl_c,
                "dt1": d1, "dt2": d2, "dt3": d3,
            })

        def _avg2(lst):
            return round(sum(lst)/len(lst), 1) if lst else None

        avg_delta = round(sum(delta_citas)/len(delta_citas), 1) if delta_citas else None

        tendencia = [
            {"mes": m,
             "total":     v["total"],
             "cumpl":     v["cumpl"],
             "no_cumpl":  v["no_cumpl"],
             "externo":   v["externo"],
             "obs_dist":  len(v["obs_set"]),
             "mans":      v["mans"],
             "rows_tot":  v["rows_tot"],
             "pct":       round(v["cumpl"] / v["total"] * 100, 1) if v["total"] else 0,
             "dt1_avg":   round(sum(v["dt1s"])/len(v["dt1s"]),1) if v.get("dt1s") else None,
             "dt2_avg":   round(sum(v["dt2s"])/len(v["dt2s"]),1) if v.get("dt2s") else None,
             "dt3_avg":   round(sum(v["dt3s"])/len(v["dt3s"]),1) if v.get("dt3s") else None}
            for m, v in sorted(tendencia_raw.items())
        ]

        result["llenos"] = {
            "total": len(ll), "exitosos": exitosos, "fallidos": fallidos,
            "cumpl_cita": cumpl_cita, "no_cumpl": no_cumpl,
            "externo_ok": externo_ok, "reprog_ajover_ok": reprog_ajover_ok, "sin_fecha_cita": sin_fecha_c,
            "motivos": dict(sorted(motivos.items(), key=lambda x: -x[1])[:20]),
            "no_cumpl_motivos": dict(sorted(no_cumpl_motivos.items(), key=lambda x: -x[1])),
            "motivos_repr": dict(sorted(motivos_repr.items(), key=lambda x: -x[1])),
            "tendencia": tendencia,
            "estados": estados, "rows": ll_rows,
            "prom": {
                "retraso_planta":   _avg2(dt1s),
                "cargue_planta":    _avg2(dt2s),
                "transito_puerto":  _avg2(dt3s),
                "delta_cita":       avg_delta,
            }
        }
    except Exception as e:
        import traceback
        stats["avisos"].append(f"Ajover LLENOS error: {e} | {traceback.format_exc()[:300]}")
        result["llenos"] = None

    return result


def manifiestos_samaritima(stats):
    """Lee la hoja VACIOS AJOVER y devuelve el set de 'Manifiesto vacío' cuyo Patio es Samaritima."""
    ruta = archivo_ajover()
    if not ruta or not os.path.isfile(ruta):
        stats["avisos"].append(f"Cruce Samaritima omitido: no encontré el archivo Ajover. (Ajusta ARCHIVO_AJOVER en el código.)")
        return set()
    try:
        ruta = _copiar_ajover(ruta)
    except Exception as e:
        stats["avisos"].append(f"Cruce Samaritima omitido: no se pudo copiar Ajover: {e}")
        return set()
    try:
        va = pd.read_excel(ruta, sheet_name=HOJA_AJOVER)
    except Exception as e:
        stats["avisos"].append(f"Cruce Samaritima omitido: no pude leer la hoja '{HOJA_AJOVER}' ({e}).")
        return set()
    va.columns = [str(c).strip() for c in va.columns]
    nmap = {_norm(c): c for c in va.columns}
    col_patio = nmap.get("patio") or next((v for k, v in nmap.items() if k == "patio"), None)
    col_man = nmap.get("manifiesto vacio") or next((v for k, v in nmap.items() if "manifiesto vacio" in k and "radicado" not in k), None)
    if not col_patio or not col_man:
        stats["avisos"].append("Cruce Samaritima omitido: no hallé las columnas 'Patio' y/o 'Manifiesto vacío' en VACIOS AJOVER.")
        return set()
    pv = va[col_patio].astype(str).str.strip().str.upper()
    objetivo = {p.upper() for p in PATIOS_SAMARITIMA}
    mask = pv.isin(objetivo)
    sam = {n for n in (env_num(v) for v in va.loc[mask, col_man].dropna()) if n}
    stats["sam_info"] = {"filas": int(mask.sum()), "manifiestos": len(sam)}
    return sam


class Cols:
    """Resuelve columnas por nombre flexible (ignora tildes/mayúsculas/espacios y acepta alias)."""
    def __init__(self, df, tipo, archivo, avisos):
        self.df, self.tipo, self.archivo, self.avisos = df, tipo, archivo, avisos
        self.map = {}
        for c in df.columns:
            self.map.setdefault(_norm(c), c)

    def tiene(self, nombre):
        return _norm(nombre) in self.map

    def serie(self, *alias, requerido=False, defecto=pd.NA, aviso=None):
        for a in alias:
            n = _norm(a)
            if n in self.map:
                return self.df[self.map[n]]
        if requerido:
            cols = "\n  ".join(repr(str(c)) for c in self.df.columns)
            raise SystemExit(
                f"\n[ERROR] En el archivo '{self.archivo}' ({self.tipo}) no encontré la columna [{alias[0]}].\n"
                f"Columnas disponibles en ese archivo:\n  {cols}\n\n"
                f"Pásame esta lista y te ajusto el nombre en el código.")
        if aviso:
            self.avisos.append(f"{self.tipo} ({self.archivo}): {aviso}")
        return pd.Series([defecto] * len(self.df), index=self.df.index, dtype="object")


# ---------------------------------------------------------------- lectura
SIG = {"orden base": "COMEX", "contable (man)": "CEDIS", "cuenta contable": "NACIONAL"}


def leer_archivo(path):
    """Recorre todas las hojas, encuentra la del encabezado correcto y detecta el tipo."""
    try:
        xls = pd.ExcelFile(path)
    except Exception:
        return None, None
    for sheet in xls.sheet_names:
        try:
            raw = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=12)
        except Exception:
            continue
        hrow = None
        for i in range(len(raw)):
            vals = {_norm(x) for x in raw.iloc[i].tolist() if pd.notna(x)}
            if vals & set(SIG):
                hrow = i
                break
        if hrow is None:
            continue
        df = pd.read_excel(xls, sheet_name=sheet, header=hrow)
        df.columns = [str(c).strip() for c in df.columns]
        nc = {_norm(c) for c in df.columns}
        for sig, tipo in SIG.items():
            if sig in nc:
                return tipo, df
    return None, None


# ---------------------------------------------------------------- limpieza por archivo
def limpiar_comex(df, stats, archivo):
    nmap = {_norm(c): c for c in df.columns}
    n0 = len(df)
    for c in ["¿Orden Borrada?", "Venta Borrada?", "¿Compra Borrada?", "Orden Borrada", "Venta Borrada"]:
        if _norm(c) in nmap:
            df = df[pd.to_numeric(df[nmap[_norm(c)]], errors="coerce").fillna(0) != 1]
    stats["borradas"] += n0 - len(df)

    C = Cols(df, "COMEX", archivo, stats["avisos"])
    fecha = pd.to_datetime(C.serie("Fecha Creacion", "Fecha Creación", "Fecha de Creacion", "Fecha",
                                   aviso="no encontré columna de fecha; el mes saldrá vacío"), errors="coerce")
    af = pd.to_numeric(C.serie("A Facturar($)", "A Facturar ($)", "A Facturar", "Valor a Facturar",
                               requerido=True).astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
    ap = pd.to_numeric(C.serie("A Pagar($)", "A Pagar ($)", "A Pagar", "Valor a Pagar",
                               requerido=True).astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
    oper = C.serie("Operacion", "Operación").astype("string")
    fuente = np.where(oper.astype(str).str.upper().str.contains("EXPO"), "EXPO", "IMPO")
    return pd.DataFrame({
        "Fuente": fuente, "EsComex": True, "Operacion": oper,
        "Segmento": np.where(fuente == "EXPO", "Comex Expo", "Comex Impo"),
        "OB": C.serie("Orden Base", requerido=True).astype("string"),
        "Manifiesto": C.serie("Envio(compra)", "Envio (compra)", "Envío(compra)", "Envio Compra").astype("string"),
        "EnvioVenta": C.serie("Envio(Venta)", "Envio (Venta)", "Envío(Venta)", "Envio Venta").astype("string"),
        "Contenedor": C.serie("Container", "Contenedor", "Contenedor (Container)", "Nº Contenedor").astype("string"),
        "Fecha": fecha.values,
        "ClienteNIT": C.serie("Cliente").astype("string"), "ClienteNombre": C.serie("Cliente Nombre").astype("string"),
        "ProvNIT": C.serie("Proveedor").astype("string"), "ProvNombre": C.serie("Proveedor Nombre").astype("string"),
        "Propiedad": np.where(C.serie("Proveedor").map(norm_nit) == NIT_PROPIO, "PROPIO", "TERCERO"),
        "Tipologia": C.serie("Tipologia", "Tipología", defecto="(Sin tipología)").fillna("(Sin tipología)").astype("string"),
        "CuentaContable": "", "Token": "TL", "Ciudad": C.serie("Agencia").astype("string"), "AltoCubicaje": False,
        "Placa": C.serie("Placa").astype("string"), "Origen": C.serie("Ciudad Origen", "Origen").astype("string"),
        "Destino": C.serie("Ciudad Destino", "Destino").astype("string"),
        "AFacturar": af.values, "APagar": ap.values,
    })


def limpiar_nacional(df, stats, archivo):
    # Solo Nacional COMEX: únicamente filas con token TL en cuenta contable
    _nmap = {_norm(c): c for c in df.columns}
    _cc = _nmap.get("cuenta contable") or _nmap.get("cuenta")
    if _cc:
        _tok = df[_cc].astype(str).map(lambda x: split_cuenta(x)[1]).str.upper()
        df = df[_tok == "TL"].reset_index(drop=True)
    C = Cols(df, "NACIONAL", archivo, stats["avisos"])
    fecha = pd.to_datetime(C.serie("Fecha Creacion", "Fecha Creación", "Fecha de Creacion", "Fecha",
                                   aviso="no encontré columna de fecha; el mes saldrá vacío"), errors="coerce")
    af = pd.to_numeric(C.serie("Envio venta (Total)", "Envío venta (Total)", "Envio Venta (Total)",
                               "Venta (Total)", "Total Venta", "Valor Venta", "A Facturar",
                               requerido=True), errors="coerce").fillna(0)
    ap = pd.to_numeric(C.serie("Envio compra (Total)", "Envío compra (Total)", "Compra (Total)",
                               "Total Compra", "A Pagar", requerido=True), errors="coerce").fillna(0)
    cuenta = C.serie("Cuenta Contable", "Cuenta contable", "Cuenta", requerido=True).astype("string")
    sc = cuenta.map(split_cuenta); ciudad = sc.map(lambda t: t[0]); token = sc.map(lambda t: t[1])
    # Reclasificar filas según Token de CuentaContable
    fuente_col = np.where(token == "CED", "CEDIS",
                 np.where(token == "TL",  "NAL-TL", "NACIONAL"))
    oper_col   = fuente_col  # Operacion = Fuente
    es_comex   = (token == "TL")
    return pd.DataFrame({
        "Fuente": fuente_col, "EsComex": es_comex, "Operacion": oper_col,
        "Segmento": [seg_from("NACIONAL", t) for t in token],
        "OB": pd.NA, "Manifiesto": C.serie("Envio", "Envío", "Manifiesto").astype("string"), "Fecha": fecha.values,
        "ClienteNIT": C.serie("Doc Cliente", "Documento Cliente", "NIT Cliente").astype("string"),
        "ClienteNombre": C.serie("Cliente", "Nombre Cliente").astype("string"),
        "ProvNIT": C.serie("Doc Afiliado", "Documento Afiliado", "NIT Afiliado").astype("string"),
        "ProvNombre": C.serie("Afiliado", "Nombre Afiliado").astype("string"),
        "Propiedad": np.where(C.serie("Doc Afiliado", "Documento Afiliado", "NIT Afiliado").map(norm_nit) == NIT_PROPIO, "PROPIO", "TERCERO"),
        "Tipologia": C.serie("Tipologia", "Tipología", defecto="(Sin tipología)").fillna("(Sin tipología)").astype("string"),
        "CuentaContable": cuenta, "Token": token.astype("string"), "Ciudad": ciudad.astype("string"),
        "AltoCubicaje": (token == "AC").values,
        "CodCliente": C.serie("Cod Cliente", "Codigo Cliente", "Código Cliente", defecto="").fillna("").astype("string"),
        "Placa": C.serie("Placa").astype("string"), "Origen": C.serie("Ciudad Origen", "Origen").astype("string"),
        "Destino": C.serie("Ciudad Destino", "Destino").astype("string"),
        "AFacturar": af.values, "APagar": ap.values,
    })


def limpiar_cedis(df, stats, archivo):
    C = Cols(df, "CEDIS", archivo, stats["avisos"])
    fecha = pd.to_datetime(C.serie("Creacion (Man)", "Creación (Man)", "Fecha Creacion", "Fecha (Man)", "Fecha",
                                   aviso="no encontré columna de fecha; el mes saldrá vacío"), errors="coerce")
    af = pd.to_numeric(C.serie("Costo Total (Venta)", "Costo total (Venta)", "Total (Venta)", "Valor Venta",
                               "A Facturar", "Envio venta (Total)", requerido=True), errors="coerce").fillna(0)
    ap = pd.to_numeric(C.serie("Costo Total (Man)", "Costo total (Man)", "Total (Man)", "A Pagar",
                               "Envio compra (Total)", requerido=True), errors="coerce").fillna(0)
    cont = C.serie("Contable (Man)", "Contable", "Cuenta Contable", requerido=True).astype("string")
    sc = cont.map(split_cuenta); ciudad = sc.map(lambda t: t[0]); token = sc.map(lambda t: t[1])
    return pd.DataFrame({
        "Fuente": "CEDIS", "EsComex": False, "Operacion": "CEDIS", "Segmento": "Cedis",
        "OB": pd.NA, "Manifiesto": C.serie("Manifiesto", "Envio").astype("string"), "Fecha": fecha.values,
        "ClienteNIT": C.serie("NIT (Cliente)", "Nit (Cliente)").astype("string"),
        "ClienteNombre": C.serie("Nombre (Cliente)").astype("string"),
        "ProvNIT": C.serie("Doc Afiliado").astype("string"),
        "ProvNombre": C.serie("Nom Afiliado", "Nombre Afiliado", "Afiliado").astype("string"),
        "Propiedad": np.where(C.serie("Doc Afiliado").map(norm_nit) == NIT_PROPIO, "PROPIO", "TERCERO"),
        "Tipologia": "(Sin tipología)",
        "CuentaContable": cont, "Token": token.astype("string"), "Ciudad": ciudad.astype("string"), "AltoCubicaje": False,
        "Placa": C.serie("Placa (Veh)", "Placa").astype("string"),
        "Origen": C.serie("Origen (Man)", "Origen", "Ciudad Origen").astype("string"),
        "Destino": C.serie("Destino (Man)", "Destino", "Ciudad Destino").astype("string"),
        "AFacturar": af.values, "APagar": ap.values,
    })


LIMPIADORES = {"COMEX": limpiar_comex, "NACIONAL": limpiar_nacional, "CEDIS": limpiar_cedis}


# ---------------------------------------------------------------- DITAR enrichment
def leer_ditar_despachos(stats):
    """Lee ditar_despachos.json exportado desde despachoscomex.tractocar.com.
    Retorna:
      lookup  dict: ManifiestoNum -> {SubOperacion, PendienteFacturar, PendientePagar}
      roundtrip_conts  set de contenedores que aparecen en IMPO y EXPO (= RoundTrip)
    """
    ruta = os.path.join(BASE, ARCHIVO_DITAR)
    if not os.path.isfile(ruta):
        stats["avisos"].append(
            f"Enriquecimiento DITAR omitido: falta '{ARCHIVO_DITAR}'. "
            "Exporta el archivo desde despachoscomex.tractocar.com y colócalo junto a procesar.py.")
        return {}, set()
    try:
        with open(ruta, encoding="utf-8") as f:
            rows = json.load(f)
    except Exception as e:
        stats["avisos"].append(f"Enriquecimiento DITAR omitido: no pude leer '{ARCHIVO_DITAR}' ({e}).")
        return {}, set()

    lookup = {}
    contenedores_impo = set()
    contenedores_expo = set()
    traslados = 0

    for r in rows:
        man_num = env_num(r.get("Manifiesto lleno", ""))
        cont = str(r.get("Contenedor", "")).strip()
        sub_op = str(r.get("Sub-operación", "")).strip()
        pend = r.get("_fact_pend") or []
        pend_fac = "Flete a facturar" in pend
        pend_pag = any("pagar" in p.lower() for p in pend)

        if cont:
            if sub_op == "Importación":
                contenedores_impo.add(cont)
            elif sub_op == "Exportación":
                contenedores_expo.add(cont)
        if sub_op == "Traslado vacíos":
            traslados += 1
        notas = str(r.get("Notas", "") or "").strip()
        if man_num:
            lookup[man_num] = {
                "SubOperacion": sub_op,
                "PendienteFacturar": pend_fac,
                "PendientePagar": pend_pag,
                "NotasWebApp": notas,
            }

    roundtrip_conts = contenedores_impo & contenedores_expo
    stats["ditar_info"] = {
        "registros": len(rows),
        "roundtrips_contenedores": len(roundtrip_conts),
        "traslados": traslados,
    }
    return lookup, roundtrip_conts


# ---------------------------------------------------------------- Excel
def escribir_excel(U, ruta):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    orden = ["Fuente", "EsComex", "Operacion", "Segmento", "SubOperacion", "EsRoundTrip", "RTDetectado",
             "OB", "Manifiesto", "EnvioVenta", "Contenedor", "CodCliente", "Fecha", "Mes", "Anio",
             "Dia", "ClienteNIT", "ClienteNombre", "ProvNIT", "ProvNombre", "Propiedad", "Tipologia",
             "CuentaContable", "Token", "Ciudad", "AltoCubicaje", "Placa", "Origen", "Destino",
             "TocoSamaritima", "AFacturar", "APagar", "Utilidad", "Margen",
             "AlertaValorDitar", "PendienteFacturar", "PendientePagar",
             "VentaTotalOB", "PagarTotalOB", "UtilidadTotalOB", "MovimientosOB",
             "MesesOB", "MultiMesOB", "SamMovs", "AjusteSamaritima"]
    cols = [c for c in orden if c in U.columns]
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        U[cols].to_excel(w, index=False, sheet_name="Union")
    wb = openpyxl.load_workbook(ruta); ws = wb["Union"]
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor="0C4A5A")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
    for col in ws.columns:
        letter = col[0].column_letter
        vals = [len(str(c.value)) for c in col[:300] if c.value is not None]
        ws.column_dimensions[letter].width = max(11, min(34, (max(vals) if vals else 10) + 2))
    last = ws.max_row
    for name in ["AFacturar", "APagar", "Utilidad", "VentaTotalOB", "PagarTotalOB", "UtilidadTotalOB", "AjusteSamaritima"]:
        if name in hdr:
            for r in range(2, last + 1):
                ws.cell(row=r, column=hdr[name]).number_format = "#,##0"
    if "Margen" in hdr:
        for r in range(2, last + 1):
            ws.cell(row=r, column=hdr["Margen"]).number_format = "0.0%"
    if "Fecha" in hdr:
        for r in range(2, last + 1):
            ws.cell(row=r, column=hdr["Fecha"]).number_format = "yyyy-mm-dd"
    wb.save(ruta)


# ---------------------------------------------------------------- obtener_union
def obtener_union(verbose=True):
    """Retorna el DataFrame unificado en memoria sin escribir a Excel."""
    carps = carpetas_datos()
    archivos = []
    for carp in carps:
        archivos += [f for f in glob.glob(os.path.join(carp, "**", "*.xls*"), recursive=True)
                     if not os.path.basename(f).startswith("~$")]
    archivos = sorted(set(archivos))
    if not archivos:
        raise RuntimeError("No encontre archivos .xlsx en ninguna carpeta.")
    stats = {"borradas": 0, "avisos": [], "sam_info": None}
    frames = []
    carp0 = carps[0]
    for path in sorted(archivos):
        try:
            tipo, df = leer_archivo(path)
        except Exception as e:
            if verbose: print(f"  · {os.path.relpath(path, carp0)}: omitido ({e})")
            continue
        if tipo not in LIMPIADORES:
            continue
        # Si el archivo viene de una carpeta cedis pero fue detectado como NACIONAL, forzar CEDIS
        if tipo == "NACIONAL" and "cedis" in path.lower():
            tipo = "CEDIS"
        rel = os.path.relpath(path, carp0)
        limpio = LIMPIADORES[tipo](df, stats, rel)
        frames.append(limpio)
        if verbose:
            print(f"  · {rel:44s} -> {tipo:8s} {len(limpio):>6,} filas")
    if not frames:
        raise RuntimeError("Ningun archivo reconocido (IMPO/EXPO/NACIONAL/CEDIS).")
    U = pd.concat(frames, ignore_index=True)
    U["Utilidad"] = U["AFacturar"] - U["APagar"]
    U["Margen"]   = np.where(U["AFacturar"] != 0, U["Utilidad"] / U["AFacturar"], 0.0)
    U["Mes"]      = U["Fecha"].dt.strftime("%Y-%m").fillna("(sin fecha)")
    U["Anio"]     = U["Fecha"].dt.year
    U["Dia"]      = U["Fecha"].dt.day
    U["FechaISO"] = U["Fecha"].dt.strftime("%Y-%m-%d")
    return U


# ---------------------------------------------------------------- main
def main():
    carps = carpetas_datos()
    print("=" * 64)
    print("TRACTOCAR · Control de Ventas — procesando")
    for c in carps:
        print("Carpeta:", c)
    if not carps:
        print("\n[ERROR] No encuentro ninguna carpeta. Revisa las variables CARPETA y CARPETA2.")
        sys.exit(1)

    archivos = []
    for carp in carps:
        archivos += [f for f in glob.glob(os.path.join(carp, "**", "*.xls*"), recursive=True)
                     if not os.path.basename(f).startswith("~$")]
    archivos = sorted(set(archivos))
    carp = carps[0]   # carpeta base para rutas relativas en el log
    if not archivos:
        print("\n[ERROR] No encontré archivos .xlsx en ninguna carpeta.")
        sys.exit(1)

    stats = {"borradas": 0, "avisos": [], "sam_info": None}
    frames, resumen = [], []
    for path in sorted(archivos):
        try:
            tipo, df = leer_archivo(path)
        except Exception as e:
            print(f"  · {os.path.relpath(path, carp)}: no se pudo leer ({e})")
            continue
        if tipo not in LIMPIADORES:
            print(f"  · {os.path.relpath(path, carp)}: tipo no reconocido, se omite")
            continue
        rel = os.path.relpath(path, carp)
        limpio = LIMPIADORES[tipo](df, stats, rel)
        frames.append(limpio)
        resumen.append({"tipo": tipo, "nombre": rel, "filas": int(len(limpio))})
        print(f"  · {rel:44s} -> {tipo:8s} {len(limpio):>6,} filas")

    if not frames:
        print("\n[ERROR] Ningún archivo reconocido (IMPO / EXPO / NACIONAL / CEDIS).")
        sys.exit(1)

    U = pd.concat(frames, ignore_index=True)
    U["Utilidad"] = U["AFacturar"] - U["APagar"]
    U["Margen"] = np.where(U["AFacturar"] != 0, U["Utilidad"] / U["AFacturar"], 0.0)
    U["Mes"] = U["Fecha"].dt.strftime("%Y-%m").fillna("(sin fecha)")
    U["Anio"] = U["Fecha"].dt.year
    U["Dia"] = U["Fecha"].dt.day
    U["FechaISO"] = U["Fecha"].dt.strftime("%Y-%m-%d")

    comex = U[(U["EsComex"]) & (U["OB"].notna())].copy()
    ob = comex.groupby("OB").agg(
        VentaTotalOB=("AFacturar", "sum"), PagarTotalOB=("APagar", "sum"),
        MovimientosOB=("OB", "size"), MesesOB=("Mes", "nunique")).reset_index()
    ob["UtilidadTotalOB"] = ob["VentaTotalOB"] - ob["PagarTotalOB"]
    ob["MultiMesOB"] = ob["MesesOB"] > 1

    # ----- listas por OB (separadas por coma): envíos de venta, manifiestos (compra), contenedores
    def _lista_por_ob(df, col):
        s = df.copy()
        s["_v"] = s[col].fillna("").astype(str).str.strip()
        s = s[~s["_v"].isin(["", "<NA>", "nan", "None", "NaN"])]
        return (s.groupby("OB")["_v"]
                 .apply(lambda x: ", ".join(sorted({str(v).strip() for v in x if str(v).strip() not in ("", "nan", "None", "<NA>", "NaN")}))))
    for col, nombre in [("EnvioVenta", "EnviosVenta"), ("Manifiesto", "Manifiestos"), ("Contenedor", "Contenedores")]:
        if col in comex.columns:
            ob = ob.merge(_lista_por_ob(comex, col).rename(nombre), on="OB", how="left")
        if nombre not in ob.columns:
            ob[nombre] = ""
        ob[nombre] = ob[nombre].fillna("")
    if "EnviosVenta" not in ob.columns:
        ob["EnviosVenta"] = ""
    ob["EnviosVenta"] = ob["EnviosVenta"].fillna("")

    # ----- cruce Samaritima: + AJUSTE por cada manifiesto vacío que tocó Samaritima
    sam = manifiestos_samaritima(stats)
    U["TocoSamaritima"] = False
    ob["SamMovs"] = 0
    ob["AjusteSamaritima"] = 0
    if sam:
        nc = U["Manifiesto"].map(env_num)
        nv = U["EnvioVenta"].map(env_num) if "EnvioVenta" in U.columns else pd.Series([None] * len(U), index=U.index)
        en_c = nc.map(lambda x: x in sam if x is not None else False)
        en_v = nv.map(lambda x: x in sam if x is not None else False)
        U["TocoSamaritima"] = U["EsComex"] & (en_c | en_v)
        matched = pd.Series([None] * len(U), index=U.index, dtype="object")
        matched[en_c.values] = nc[en_c.values]
        falta = matched.isna() & en_v
        matched[falta.values] = nv[falta.values]
        U["_SamManif"] = matched
        c2 = U[(U["EsComex"]) & (U["OB"].notna()) & (U["_SamManif"].notna())]
        if len(c2):
            sm = c2.groupby("OB")["_SamManif"].nunique().rename("SamMovs").reset_index()
            ob = ob.drop(columns=["SamMovs"]).merge(sm, on="OB", how="left")
            ob["SamMovs"] = ob["SamMovs"].fillna(0).astype(int)
            ob["AjusteSamaritima"] = ob["SamMovs"] * AJUSTE_SAMARITIMA
        U = U.drop(columns=["_SamManif"])

    U = U.merge(ob[["OB", "VentaTotalOB", "PagarTotalOB", "UtilidadTotalOB", "MovimientosOB",
                    "MesesOB", "MultiMesOB", "SamMovs", "AjusteSamaritima"]], on="OB", how="left")

    # ----- Enriquecimiento DITAR: Sub-operación, RoundTrip, pendientes, alertas de valor
    ditar_lookup, _ = leer_ditar_despachos(stats)   # web app → sub-operación y pendientes
    U["SubOperacion"] = pd.NA
    U["PendienteFacturar"] = False
    U["PendientePagar"] = False
    U["EsRoundTrip"] = False
    U["RTContenedor"] = False  # RT confirmado: mismo contenedor en IMPO y EXPO
    U["NotasWebApp"] = ""
    U["AlertaValorDitar"] = False

    es_ditar = U["ClienteNombre"].astype(str).str.upper().str.contains(DITAR_NOMBRE_PATRON, na=False)

    # Sub-operación, pendientes y notas: cruce por Manifiesto (web app → archivos de venta)
    if ditar_lookup:
        man_nums = U["Manifiesto"].map(env_num)
        U["SubOperacion"] = man_nums.map(lambda x: ditar_lookup[x]["SubOperacion"] if x and x in ditar_lookup else pd.NA)
        U["PendienteFacturar"] = man_nums.map(lambda x: ditar_lookup[x]["PendienteFacturar"] if x and x in ditar_lookup else False)
        U["PendientePagar"] = man_nums.map(lambda x: ditar_lookup[x]["PendientePagar"] if x and x in ditar_lookup else False)
        U["NotasWebApp"] = man_nums.map(lambda x: ditar_lookup[x]["NotasWebApp"] if x and x in ditar_lookup else "")

    # EsRoundTrip: OB contiene "RT" en su nombre (detección desde archivos de venta)
    mask_rt_ob = U["EsComex"] & es_ditar & U["OB"].astype(str).str.upper().str.contains(r"\bRT\b", na=False)
    U.loc[mask_rt_ob, "EsRoundTrip"] = True
    U.loc[mask_rt_ob & U["SubOperacion"].isna(), "SubOperacion"] = "RoundTrip"

    # RTContenedor: mismo contenedor aparece en IMPO y EXPO (confirmación adicional)
    cont_ok = U["Contenedor"].astype(str).str.strip()
    ditar_impo_conts = set(cont_ok[(U["EsComex"]) & (U["Fuente"] == "IMPO") & es_ditar & (cont_ok != "") & (cont_ok != "nan") & (cont_ok != "<NA>")])
    ditar_expo_conts = set(cont_ok[(U["EsComex"]) & (U["Fuente"] == "EXPO") & es_ditar & (cont_ok != "") & (cont_ok != "nan") & (cont_ok != "<NA>")])
    roundtrip_conts_excel = ditar_impo_conts & ditar_expo_conts
    if roundtrip_conts_excel:
        U["RTContenedor"] = U["EsComex"] & es_ditar & cont_ok.isin(roundtrip_conts_excel)

    # Traslados: OB contiene la palabra TRASLADO (desde los archivos de venta)
    mask_traslado = U["EsComex"] & U["OB"].astype(str).str.upper().str.contains("TRASLADO", na=False)
    U.loc[mask_traslado, "SubOperacion"] = "Traslado"

    stats["ditar_rt_conts"] = len(roundtrip_conts_excel)
    stats["ditar_traslados_ob"] = int(mask_traslado.sum())

    U["AlertaValorDitar"] = es_ditar & ((U["AFacturar"] > DITAR_ALERTA_MONTO) | (U["APagar"] > DITAR_ALERTA_MONTO))

    ruta_xlsx = os.path.join(BASE, "TRACTOCAR_UNIFICADO.xlsx")
    escribir_excel(U, ruta_xlsx)

    vista = ["Fuente", "EsComex", "Segmento", "SubOperacion", "EsRoundTrip", "RTContenedor",
             "OB", "Manifiesto", "EnvioVenta", "Contenedor", "Mes", "FechaISO", "ClienteNIT",
             "ClienteNombre", "ProvNIT", "ProvNombre", "Propiedad", "Tipologia", "AltoCubicaje",
             "Placa", "Origen", "Destino", "AFacturar", "APagar", "Utilidad", "Margen",
             "AlertaValorDitar", "PendienteFacturar", "PendientePagar", "TocoSamaritima",
             "VentaTotalOB", "MesesOB", "NotasWebApp"]
    d = U[[c for c in vista if c in U.columns]].copy()
    for c in ["AFacturar", "APagar", "Utilidad"]:
        d[c] = d[c].round(2)
    d["Margen"] = d["Margen"].round(4)
    records = json.loads(d.to_json(orient="records"))

    def _obt_row(r):
        return [round(float(r["VentaTotalOB"]), 2), round(float(r["PagarTotalOB"]), 2),
                int(r["MovimientosOB"]), int(r["MesesOB"]),
                (r["EnviosVenta"] if isinstance(r["EnviosVenta"], str) else ""),
                int(r["SamMovs"]), round(float(r["AjusteSamaritima"]), 2),
                (r["Manifiestos"] if isinstance(r["Manifiestos"], str) else ""),
                (r["Contenedores"] if isinstance(r["Contenedores"], str) else "")]
    obt = {str(r["OB"]): _obt_row(r) for _, r in ob.iterrows()}
    meta = {"generado": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "meses": sorted(U["Mes"].unique().tolist()),
            "movimientos": int(len(U)), "archivos": resumen, "borradas": int(stats["borradas"]),
            "samaritima": {"obs": int((ob["SamMovs"] > 0).sum()),
                           "ajuste": float(ob["AjusteSamaritima"].sum()),
                           "info": stats.get("sam_info")}}

    ajcomex_data = leer_ajover_comex(stats)
    ajover_data = leer_ajover_completo(stats)
    if ajover_data:
        v = ajover_data.get("vacios") or {}
        l = ajover_data.get("llenos") or {}
        print(f"  Ajover VACIOS              : {v.get('total',0)} filas | "
              f"A tiempo: {v.get('a_tiempo',0)} | Tarde: {v.get('tarde',0)}")
        print(f"  Ajover LLENOS              : {l.get('total',0)} filas | "
              f"Exitosos: {l.get('exitosos',0)} | Fallidos: {l.get('fallidos',0)}")

    payload = "window.DATA=%s;window.OBT=%s;window.META=%s;window.AJOVER=%s;window.AJCOMEX=%s;" % (
        json.dumps(records, ensure_ascii=False),
        json.dumps(obt, ensure_ascii=False),
        json.dumps(meta, ensure_ascii=False),
        json.dumps(ajover_data or {}, ensure_ascii=False),
        json.dumps(ajcomex_data or {}, ensure_ascii=False))
    payload = payload.replace("</", "<\\/")

    tpl_path = os.path.join(BASE, "plantilla.html")
    if not os.path.exists(tpl_path):
        print("\n[ERROR] Falta 'plantilla.html' en la misma carpeta que este script.")
        sys.exit(1)
    tpl = open(tpl_path, encoding="utf-8").read()
    ruta_html = os.path.join(BASE, "index.html")
    open(ruta_html, "w", encoding="utf-8").write(tpl.replace("/*__DATOS__*/", payload))

    af, ap = U["AFacturar"].sum(), U["APagar"].sum()
    print("-" * 64)
    print(f"  Registros borrados excluidos : {stats['borradas']}")
    print(f"  Movimientos (unión)          : {len(U):,}")
    print(f"  Viajes únicos                : {U['Manifiesto'].nunique():,}")
    print(f"  OB de COMEX                  : {comex['OB'].nunique():,}")
    print(f"  A facturar                   : ${af:,.0f}")
    print(f"  A pagar                      : ${ap:,.0f}")
    print(f"  Utilidad                     : ${af-ap:,.0f}  (margen {(af-ap)/af*100:.1f}%)" if af else "  Utilidad : -")
    si = stats.get("sam_info")
    if si:
        print(f"  Cruce Samaritima             : {si['manifiestos']} manifiestos ({si['filas']} filas patio) | "
              f"OB con ajuste: {(ob['SamMovs']>0).sum()} | ajuste total: ${ob['AjusteSamaritima'].sum():,.0f}")
    di = stats.get("ditar_info")
    if di:
        print(f"  DITAR web app                : {di['registros']} registros | Traslados vacíos: {di['traslados']}")
    rt_conts = stats.get("ditar_rt_conts", 0)
    alertas = int(U["AlertaValorDitar"].sum())
    pend_f = int(U["PendienteFacturar"].sum())
    pend_p = int(U["PendientePagar"].sum())
    rt_filas = int(U["EsRoundTrip"].sum())
    traslados_ob = stats.get("ditar_traslados_ob", 0)
    print(f"  DITAR RoundTrip              : {rt_conts} contenedores únicos | {rt_filas} filas")
    print(f"  DITAR Traslados (OB)         : {traslados_ob} filas con 'TRASLADO' en OB")
    print(f"  DITAR alertas valor          : {alertas} filas > ${DITAR_ALERTA_MONTO:,} | "
          f"Pend. facturar: {pend_f} | Pend. pagar: {pend_p}")
    if stats["avisos"]:
        print("  Avisos:")
        for a in stats["avisos"]:
            print("   -", a)
    print("-" * 64)
    print(f"  Excel    -> {ruta_xlsx}")
    print(f"  Tablero  -> {ruta_html}")
    print("=" * 64)

    try:
        webbrowser.open("file://" + ruta_html.replace(os.sep, "/"))
    except Exception:
        pass


if __name__ == "__main__":
    main()
